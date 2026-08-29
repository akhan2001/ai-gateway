"""Export endpoint tests — no database, no network.

The three exports are the only endpoints that build a *file* rather than a JSON
body, so the thing worth pinning down is that each one produces a well-formed
document of the right type for both a busy workspace and an empty one. An empty
month is the case most likely to break a report generator, and it is also the
case every new customer hits first.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from typing import AsyncIterator
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient

from src import db as dbmod
from src.auth import current_workspace
from src.main import app

WORKSPACE = uuid4()
ANCHOR = datetime(2026, 8, 15, tzinfo=timezone.utc)

BY_MODEL = [
    {
        "model_id": "openai/gpt-4o",
        "provider": "openai",
        "requests": 1200,
        "input_tokens": 4_000_000,
        "output_tokens": 900_000,
        "cost_usd": Decimal("142.50"),
        "acpi_bench_usd": Decimal("96.20"),
        "overpay_usd": Decimal("46.30"),
    },
    {
        "model_id": "anthropic/claude-sonnet-4",
        "provider": "anthropic",
        "requests": 800,
        "input_tokens": 2_100_000,
        "output_tokens": 410_000,
        "cost_usd": Decimal("61.10"),
        "acpi_bench_usd": Decimal("64.90"),
        "overpay_usd": Decimal("-3.80"),
    },
    # Sub-cent totals are real — they are what a workspace's first day looks
    # like — and must survive the round trip without collapsing to zero.
    {
        "model_id": "openai/gpt-4o-mini",
        "provider": "openai",
        "requests": 5400,
        "input_tokens": 9_800_000,
        "output_tokens": 1_200_000,
        "cost_usd": Decimal("0.000007"),
        "acpi_bench_usd": Decimal("0.000093"),
        "overpay_usd": Decimal("-0.000086"),
    },
]

DAILY = [
    {
        "day": ANCHOR - timedelta(days=offset),
        "requests": 300 + offset,
        "input_tokens": 100_000,
        "output_tokens": 20_000,
        "cost_usd": Decimal(str(4 + (offset % 5))),
        "acpi_bench_usd": Decimal("3.9"),
        "overpay_usd": Decimal("0.4"),
    }
    for offset in range(13, -1, -1)
]

RAW_RECORD_COUNT = 2_000
RAW = [
    {
        "timestamp": ANCHOR - timedelta(hours=hour),
        "provider": "openai",
        "model_id": "openai/gpt-4o",
        "input_tokens": 900,
        "output_tokens": 210,
        "cost_usd": Decimal("0.00412"),
        "acpi_bench_usd": Decimal("0.00301"),
        "overpay_usd": Decimal("0.00111"),
    }
    for hour in range(RAW_RECORD_COUNT)
]


@pytest.fixture
def client(request, monkeypatch) -> TestClient:
    """A client whose database returns either the fixtures above or nothing."""
    empty = getattr(request, "param", False)

    async def fake_fetch(query: str, *args) -> list[dict]:
        if empty:
            return []
        return DAILY if "time_bucket" in query else BY_MODEL

    async def fake_iterate(query: str, *args, prefetch: int = 1_000) -> AsyncIterator[dict]:
        if empty:
            return
        for row in RAW:
            yield row

    monkeypatch.setattr(dbmod.db, "fetch", fake_fetch)
    monkeypatch.setattr(dbmod.db, "iterate", fake_iterate)
    app.dependency_overrides[current_workspace] = lambda: WORKSPACE
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()


@pytest.fixture
def empty_client(client) -> TestClient:
    return client


# ── happy paths ─────────────────────────────────────────────────────────────


def test_csv_streams_every_record(client):
    response = client.get("/api/v1/export/csv?days=30")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert 'filename="tokenix-usage-30d.csv"' in response.headers["content-disposition"]

    rows = list(csv.reader(io.StringIO(response.text)))
    assert rows[0] == [
        "date",
        "provider",
        "model_id",
        "input_tokens",
        "output_tokens",
        "cost_usd",
        "acpi_bench_usd",
        "overpay_usd",
    ]
    # Batched writes must not drop or duplicate the tail of the stream.
    assert len(rows) == RAW_RECORD_COUNT + 1
    assert rows[1][1:3] == ["openai", "openai/gpt-4o"]


def test_excel_has_three_named_sheets(client):
    from openpyxl import load_workbook

    response = client.get("/api/v1/export/excel?days=30")

    assert response.status_code == 200
    assert 'filename="tokenix-usage-30d.xlsx"' in response.headers["content-disposition"]

    book = load_workbook(io.BytesIO(response.content))
    assert book.sheetnames == ["Summary", "By Model", "Daily"]
    assert book["By Model"].max_row == len(BY_MODEL) + 1
    assert book["Daily"].max_row == len(DAILY) + 1

    # Totals are computed here, not read from a summary query, so they are
    # worth asserting against the rows they are derived from.
    expected_cost = sum(float(row["cost_usd"]) for row in BY_MODEL)
    summary = {row[0].value: row[1].value for row in book["Summary"].iter_rows(min_row=4)}
    assert summary["Total spend (USD)"] == pytest.approx(expected_cost)
    assert summary["Total requests"] == sum(int(row["requests"]) for row in BY_MODEL)


def test_pdf_is_a_pdf(client):
    response = client.get("/api/v1/export/pdf?month=2026-08")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert 'filename="tokenix-report-2026-08.pdf"' in response.headers["content-disposition"]
    assert response.content.startswith(b"%PDF")
    assert response.content.rstrip().endswith(b"%%EOF")


# ── the empty workspace ─────────────────────────────────────────────────────


@pytest.mark.parametrize("client", [True], indirect=True)
def test_exports_survive_an_empty_workspace(client):
    """A workspace with no traffic gets real files, not a 500."""
    csv_response = client.get("/api/v1/export/csv?days=30")
    assert csv_response.status_code == 200
    # Header only — an empty export is still a valid CSV with column names.
    assert csv_response.text.strip().split("\n") == [
        "date,provider,model_id,input_tokens,output_tokens,cost_usd,acpi_bench_usd,overpay_usd"
    ]

    excel_response = client.get("/api/v1/export/excel?days=30")
    assert excel_response.status_code == 200
    assert excel_response.content.startswith(b"PK")

    pdf_response = client.get("/api/v1/export/pdf?month=2026-08")
    assert pdf_response.status_code == 200
    assert pdf_response.content.startswith(b"%PDF")


# ── input validation ────────────────────────────────────────────────────────


@pytest.mark.parametrize("month", ["august", "2026-13", "2026", "2026-8", ""])
def test_pdf_rejects_a_malformed_month(client, month):
    assert client.get(f"/api/v1/export/pdf?month={month}").status_code == 400


def test_pdf_requires_a_month(client):
    assert client.get("/api/v1/export/pdf").status_code == 422


@pytest.mark.parametrize("days", [0, -1, 999])
def test_csv_rejects_an_out_of_range_window(client, days):
    assert client.get(f"/api/v1/export/csv?days={days}").status_code == 422
