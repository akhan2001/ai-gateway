"""Export endpoints — the same ledger the dashboard draws, as a file.

    GET /api/v1/export/csv?days=30      raw usage records
    GET /api/v1/export/excel?days=30    three formatted sheets
    GET /api/v1/export/pdf?month=2026-08 monthly summary report

Auth is the shared `current_workspace` dependency, so both callers work
unchanged: the CLI presents a `txk-` bearer key, the dashboard proxies with
`x-internal-token` + `x-workspace-id`. The workspace is resolved by that
dependency and never taken from a query parameter.

CSV streams through a server-side cursor because a month of raw records is
unbounded in a way the dashboard endpoints never are. Excel and PDF aggregate
first, so they stay small by construction.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any, AsyncIterator
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse

from ..auth import current_workspace
from ..db import db, f

router = APIRouter(prefix="/api/v1/export", tags=["export"])

# Rows are yielded to the client in batches rather than one at a time; a single
# `write` per record turns a large export into millions of tiny chunks.
_CSV_BATCH = 500

# Below this a percentage swing is noise, not a finding. Same threshold the
# /benchmark endpoint uses, for the same reason.
_MIN_SPEND_FOR_OPPORTUNITY_USD = 1.0

CSV_HEADERS = [
    "date",
    "provider",
    "model_id",
    "input_tokens",
    "output_tokens",
    "cost_usd",
    "acpi_bench_usd",
    "overpay_usd",
]

_RAW_RECORDS_SQL = """
    SELECT "timestamp",
           provider,
           model_id,
           input_tokens,
           output_tokens,
           cost_usd,
           acpi_bench_usd,
           overpay_usd
    FROM usage_records
    WHERE workspace_id = $1 AND priced
      AND "timestamp" >= NOW() - make_interval(days => $2)
    ORDER BY "timestamp"
"""

_BY_MODEL_SQL = """
    SELECT model_id,
           provider,
           COUNT(*)            AS requests,
           SUM(input_tokens)   AS input_tokens,
           SUM(output_tokens)  AS output_tokens,
           SUM(cost_usd)       AS cost_usd,
           SUM(acpi_bench_usd) AS acpi_bench_usd,
           SUM(overpay_usd)    AS overpay_usd
    FROM usage_records
    WHERE workspace_id = $1 AND priced
      AND "timestamp" >= {window}
    GROUP BY model_id, provider
    ORDER BY cost_usd DESC
"""

_DAILY_SQL = """
    SELECT time_bucket(INTERVAL '1 day', "timestamp") AS day,
           COUNT(*)            AS requests,
           SUM(input_tokens)   AS input_tokens,
           SUM(output_tokens)  AS output_tokens,
           SUM(cost_usd)       AS cost_usd,
           SUM(acpi_bench_usd) AS acpi_bench_usd,
           SUM(overpay_usd)    AS overpay_usd
    FROM usage_records
    WHERE workspace_id = $1 AND priced
      AND "timestamp" >= {window}
    GROUP BY day
    ORDER BY day
"""

# The two windows the exports run over. Interpolated as a fragment, never as
# user input: `days` and `month` still travel as bound parameters.
_ROLLING_WINDOW = 'NOW() - make_interval(days => $2)'
_MONTH_WINDOW = '$2::date AND "timestamp" < ($2::date + INTERVAL \'1 month\')'


def _disposition(filename: str) -> dict[str, str]:
    return {"content-disposition": f'attachment; filename="{filename}"'}


# strptime's %m accepts an unpadded month, so `2026-8` would parse here while
# the dashboard's download route rejects it. Both entry points reach the same
# report, so both enforce the same shape.
_MONTH_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


def _parse_month(month: str) -> date:
    """`2026-08` → the first of that month, or a 400."""
    if not _MONTH_RE.match(month):
        raise HTTPException(
            status_code=400, detail="month must be formatted YYYY-MM, e.g. 2026-08"
        )
    return datetime.strptime(month, "%Y-%m").date()


# ── CSV ─────────────────────────────────────────────────────────────────────


async def _csv_chunks(workspace_id: UUID, days: int) -> AsyncIterator[str]:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(CSV_HEADERS)

    pending = 0
    async for row in db.iterate(_RAW_RECORDS_SQL, workspace_id, days):
        writer.writerow(
            [
                row["timestamp"].date().isoformat(),
                row["provider"],
                row["model_id"],
                int(row["input_tokens"] or 0),
                int(row["output_tokens"] or 0),
                f"{f(row['cost_usd']):.8f}",
                f"{f(row['acpi_bench_usd']):.8f}",
                f"{f(row['overpay_usd']):.8f}",
            ]
        )
        pending += 1
        if pending >= _CSV_BATCH:
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)
            pending = 0

    # The header alone still counts as a body worth sending — an empty export
    # should arrive as a valid file with column names, not as nothing.
    tail = buffer.getvalue()
    if tail:
        yield tail


@router.get("/csv")
async def export_csv(
    days: int = Query(30, ge=1, le=365),
    workspace_id: UUID = Depends(current_workspace),
) -> StreamingResponse:
    return StreamingResponse(
        _csv_chunks(workspace_id, days),
        media_type="text/csv; charset=utf-8",
        headers=_disposition(f"tokenix-usage-{days}d.csv"),
    )


# ── shared aggregates ───────────────────────────────────────────────────────


async def _aggregates(
    workspace_id: UUID, window: str, bound: Any
) -> tuple[list[dict], list[dict]]:
    by_model = await db.fetch(_BY_MODEL_SQL.format(window=window), workspace_id, bound)
    daily = await db.fetch(_DAILY_SQL.format(window=window), workspace_id, bound)
    return by_model, daily


def _totals(by_model: list[dict]) -> dict[str, float]:
    cost = sum(f(r["cost_usd"]) for r in by_model)
    bench = sum(f(r["acpi_bench_usd"]) for r in by_model)
    return {
        "requests": sum(int(r["requests"]) for r in by_model),
        "input_tokens": sum(int(r["input_tokens"] or 0) for r in by_model),
        "output_tokens": sum(int(r["output_tokens"] or 0) for r in by_model),
        "cost_usd": cost,
        "acpi_bench_usd": bench,
        "overpay_usd": cost - bench,
        "overpay_pct": ((cost - bench) / bench * 100) if bench else None,
    }


def _opportunities(by_model: list[dict], period: str) -> list[dict]:
    out = []
    for row in by_model:
        overpay = f(row["overpay_usd"])
        bench = f(row["acpi_bench_usd"])
        if overpay <= _MIN_SPEND_FOR_OPPORTUNITY_USD or not bench:
            continue
        out.append(
            {
                "model_id": row["model_id"],
                "overpay_usd": overpay,
                "overpay_pct": overpay / bench * 100,
                "action": (
                    f"{row['model_id']} ran {overpay / bench * 100:.0f}% above the ACPI "
                    f"market rate. Moving this workload to a comparable cheaper model "
                    f"would have recovered about ${overpay:,.2f} {period}."
                ),
            }
        )
    return out[:3]


# ── Excel ───────────────────────────────────────────────────────────────────


@router.get("/excel")
async def export_excel(
    days: int = Query(30, ge=1, le=365),
    workspace_id: UUID = Depends(current_workspace),
) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Both sheets are aggregates — at most one row per day and one per model —
    # so the workbook is bounded regardless of how much traffic it summarises.
    # The CSV is the export that needs streaming; this one does not.
    by_model, daily = await _aggregates(workspace_id, _ROLLING_WINDOW, days)
    totals = _totals(by_model)
    workbook = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1C20")
    money = '"$"#,##0.000000'
    whole = "#,##0"

    def write_sheet(sheet, headers: list[str], rows: list[list], formats: dict[int, str]):
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        for row in rows:
            sheet.append(row)
        for column, fmt in formats.items():
            for cell in sheet[get_column_letter(column)][1:]:
                cell.number_format = fmt
        for index, header in enumerate(headers, start=1):
            width = max(len(str(header)) + 2, 14)
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"

    # Sheet 1 — Summary
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Tokenix usage report"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Last {days} days"
    summary["A2"].font = Font(color="565F6E")
    rows = [
        ("Total requests", totals["requests"], whole),
        ("Input tokens", totals["input_tokens"], whole),
        ("Output tokens", totals["output_tokens"], whole),
        ("Total spend (USD)", totals["cost_usd"], money),
        ("ACPI benchmark (USD)", totals["acpi_bench_usd"], money),
        ("Delta vs benchmark (USD)", totals["overpay_usd"], money),
        (
            "Delta vs benchmark (%)",
            totals["overpay_pct"] if totals["overpay_pct"] is not None else "n/a",
            "0.00",
        ),
        ("Models used", len(by_model), whole),
    ]
    for offset, (label, value, fmt) in enumerate(rows, start=4):
        summary[f"A{offset}"] = label
        summary[f"B{offset}"] = value
        if isinstance(value, (int, float)):
            summary[f"B{offset}"].number_format = fmt
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 20

    # Sheet 2 — By Model
    write_sheet(
        workbook.create_sheet("By Model"),
        [
            "Model",
            "Provider",
            "Requests",
            "Input tokens",
            "Output tokens",
            "Cost (USD)",
            "ACPI benchmark (USD)",
            "Delta (USD)",
            "Delta (%)",
        ],
        [
            [
                r["model_id"],
                r["provider"],
                int(r["requests"]),
                int(r["input_tokens"] or 0),
                int(r["output_tokens"] or 0),
                f(r["cost_usd"]),
                f(r["acpi_bench_usd"]),
                f(r["overpay_usd"]),
                (f(r["overpay_usd"]) / f(r["acpi_bench_usd"]) * 100)
                if f(r["acpi_bench_usd"])
                else None,
            ]
            for r in by_model
        ],
        {3: whole, 4: whole, 5: whole, 6: money, 7: money, 8: money, 9: "0.00"},
    )

    # Sheet 3 — Daily
    write_sheet(
        workbook.create_sheet("Daily"),
        [
            "Date",
            "Requests",
            "Input tokens",
            "Output tokens",
            "Cost (USD)",
            "ACPI benchmark (USD)",
            "Delta (USD)",
        ],
        [
            [
                r["day"].date().isoformat(),
                int(r["requests"]),
                int(r["input_tokens"] or 0),
                int(r["output_tokens"] or 0),
                f(r["cost_usd"]),
                f(r["acpi_bench_usd"]),
                f(r["overpay_usd"]),
            ]
            for r in daily
        ],
        {2: whole, 3: whole, 4: whole, 5: money, 6: money, 7: money},
    )

    stream = io.BytesIO()
    workbook.save(stream)
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_disposition(f"tokenix-usage-{days}d.xlsx"),
    )


# ── PDF ─────────────────────────────────────────────────────────────────────


@router.get("/pdf")
async def export_pdf(
    month: str = Query(..., description="Report month, formatted YYYY-MM"),
    workspace_id: UUID = Depends(current_workspace),
) -> Response:
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    start = _parse_month(month)
    by_model, daily = await _aggregates(workspace_id, _MONTH_WINDOW, start)
    totals = _totals(by_model)
    opportunities = _opportunities(by_model, f"in {start:%B %Y}")

    ink = colors.HexColor("#1a1c20")
    muted = colors.HexColor("#565f6e")
    faint = colors.HexColor("#8b8576")
    rule = colors.HexColor("#e3ddd0")
    gold = colors.HexColor("#9a7b2e")

    base = getSampleStyleSheet()
    title = ParagraphStyle(
        "TxTitle", parent=base["Title"], fontName="Times-Roman", fontSize=24,
        textColor=ink, alignment=0, spaceAfter=2,
    )
    kicker = ParagraphStyle(
        "TxKicker", parent=base["Normal"], fontName="Helvetica", fontSize=7.5,
        textColor=faint, spaceAfter=10, leading=11,
    )
    section = ParagraphStyle(
        "TxSection", parent=base["Normal"], fontName="Times-Roman", fontSize=13,
        textColor=ink, spaceBefore=18, spaceAfter=8,
    )
    body = ParagraphStyle(
        "TxBody", parent=base["Normal"], fontName="Helvetica", fontSize=9,
        textColor=muted, leading=14,
    )

    def usd(value: float) -> str:
        if abs(value) >= 1:
            return f"${value:,.2f}"
        if value == 0:
            return "$0.00"
        return f"${value:,.6f}"

    story: list = [
        Paragraph("Tokenix", title),
        Paragraph(
            f"AI SPEND REPORT &nbsp;·&nbsp; {start:%B %Y}".upper(),
            kicker,
        ),
        HRFlowable(width="100%", thickness=0.6, color=ink, spaceAfter=16),
    ]

    if not by_model:
        story += [
            Paragraph("No priced traffic this month", section),
            Paragraph(
                "No requests were recorded through the Tokenix gateway during "
                f"{start:%B %Y}. Once traffic flows, this report summarises what it "
                "cost and how that compared with the ACPI market rate.",
                body,
            ),
        ]
    else:
        delta_pct = totals["overpay_pct"]
        headline = [
            ["Total spend", usd(totals["cost_usd"])],
            ["ACPI benchmark", usd(totals["acpi_bench_usd"])],
            [
                "Versus market",
                f"{delta_pct:+.1f}%" if delta_pct is not None else "—",
            ],
            ["Requests", f"{totals['requests']:,}"],
            [
                "Tokens",
                f"{totals['input_tokens'] + totals['output_tokens']:,}",
            ],
        ]
        headline_table = Table(headline, colWidths=[70 * mm, 95 * mm])
        headline_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
                    ("FONTNAME", (1, 0), (1, -1), "Courier"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("TEXTCOLOR", (0, 0), (0, -1), muted),
                    ("TEXTCOLOR", (1, 0), (1, -1), ink),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("LINEBELOW", (0, 0), (-1, -2), 0.4, rule),
                ]
            )
        )
        story += [headline_table]

        if delta_pct is not None:
            verdict = (
                f"Spend ran {abs(delta_pct):.1f}% "
                f"{'above' if delta_pct > 0 else 'below'} the ACPI market rate this "
                f"month, a difference of {usd(abs(totals['overpay_usd']))}."
            )
            story += [Spacer(1, 12), Paragraph(verdict, body)]

        # Daily trend. Drawn with reportlab's own primitives so the report adds
        # no charting dependency for one bar chart.
        if len(daily) >= 2:
            story += [Paragraph("Daily spend", section)]
            values = [f(r["cost_usd"]) for r in daily]
            drawing = Drawing(460, 140)
            chart = VerticalBarChart()
            chart.x, chart.y = 30, 24
            chart.width, chart.height = 415, 100
            chart.data = [values]
            chart.bars[0].fillColor = gold
            chart.bars[0].strokeColor = None
            chart.valueAxis.valueMin = 0
            chart.valueAxis.strokeColor = rule
            chart.valueAxis.gridStrokeColor = rule
            chart.valueAxis.labels.fontName = "Courier"
            chart.valueAxis.labels.fontSize = 6
            chart.categoryAxis.strokeColor = rule
            chart.categoryAxis.labels.fontName = "Helvetica"
            chart.categoryAxis.labels.fontSize = 5.5
            chart.categoryAxis.labels.angle = 90
            chart.categoryAxis.labels.dy = -8
            # Every day would collide at 5.5pt; label roughly six of them.
            step = max(1, len(daily) // 6)
            chart.categoryAxis.categoryNames = [
                r["day"].strftime("%d") if i % step == 0 else ""
                for i, r in enumerate(daily)
            ]
            drawing.add(chart)
            story += [drawing]

        story += [Paragraph("Top models by spend", section)]
        top = by_model[:5]
        model_rows = [["Model", "Provider", "Requests", "Cost", "vs market"]]
        for row in top:
            bench = f(row["acpi_bench_usd"])
            pct = (f(row["overpay_usd"]) / bench * 100) if bench else None
            model_rows.append(
                [
                    row["model_id"],
                    row["provider"],
                    f"{int(row['requests']):,}",
                    usd(f(row["cost_usd"])),
                    f"{pct:+.1f}%" if pct is not None else "—",
                ]
            )
        model_table = Table(
            model_rows, colWidths=[58 * mm, 28 * mm, 24 * mm, 30 * mm, 25 * mm]
        )
        model_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 7),
                    ("TEXTCOLOR", (0, 0), (-1, 0), faint),
                    ("FONTNAME", (0, 1), (1, -1), "Helvetica"),
                    ("FONTNAME", (2, 1), (-1, -1), "Courier"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("TEXTCOLOR", (0, 1), (-1, -1), ink),
                    ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.6, ink),
                    ("LINEBELOW", (0, 1), (-1, -2), 0.4, rule),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story += [model_table]

        story += [Paragraph("Savings opportunities", section)]
        if opportunities:
            for item in opportunities:
                story += [
                    Paragraph(f"— {item['action']}", body),
                    Spacer(1, 6),
                ]
        else:
            story += [
                Paragraph(
                    "No model ran far enough above the ACPI market rate to be worth "
                    "moving this month.",
                    body,
                )
            ]

    footer = ParagraphStyle(
        "TxFooter", parent=body, fontSize=7, textColor=faint, alignment=TA_RIGHT
    )
    story += [
        Spacer(1, 24),
        HRFlowable(width="100%", thickness=0.4, color=rule, spaceAfter=6),
        Paragraph(
            "Benchmarked against the Tokenix AI Compute Price Index (ACPI). "
            "tokenixindex.com",
            footer,
        ),
    ]

    stream = io.BytesIO()
    SimpleDocTemplate(
        stream,
        pagesize=A4,
        leftMargin=22 * mm,
        rightMargin=22 * mm,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
        title=f"Tokenix AI spend report — {start:%B %Y}",
        author="Tokenix",
    ).build(story)

    return Response(
        content=stream.getvalue(),
        media_type="application/pdf",
        headers=_disposition(f"tokenix-report-{start:%Y-%m}.pdf"),
    )
